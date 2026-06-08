import os
import json
import tempfile
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from src.modulos.quadro_horario.relatorios.repository import RelatorioQuadroHorarioRepository

# ==========================================
# UTILITÁRIOS DE DADOS PARA A PESQUISA
# ==========================================
def safe_float(x):
    try:
        if x is None or x == "": return 0.0
        if isinstance(x, str): x = x.replace(",", ".")
        return float(x)
    except: return 0.0

def ensure_payload_list(payload):
    try:
        if isinstance(payload, str):
            try: payload = json.loads(payload)
            except: pass
        if not payload: return []

        # TRATAMENTO PERFEITO PARA O SEU JSON: {"datas": [...], "tabelas": [...]}
        if isinstance(payload, dict) and "tabelas" in payload and isinstance(payload["tabelas"], list):
            datas = payload.get("datas", [])
            out = []
            for i, tab in enumerate(payload["tabelas"]):
                norm = _normalize_table_static(tab)
                
                # MÁGICA DE RENOMEAÇÃO: "Resumo 1" -> "Relatório 1 - 06/08/2025"
                nome_orig = str(norm.get("nome", ""))
                data_ref = datas[i] if i < len(datas) else ""
                
                if "Resumo" in nome_orig:
                    num = nome_orig.replace("Resumo", "").strip()
                    norm["nome"] = f"Relatório {num} - {data_ref}" if data_ref else f"Relatório {num}"
                else:
                    norm["nome"] = f"{nome_orig} - {data_ref}" if data_ref else nome_orig
                    
                if norm["columns"] or norm["rows"]:
                    out.append(norm)
            if out: return out

        # FALLBACK PARA FORMATOS ANTIGOS (Caçador de Tabelas)
        tables = []
        def hunt_tables(obj, name_context="Tabela"):
            if isinstance(obj, dict):
                if any(k in obj for k in ["rows", "dados", "linhas", "columns", "colunas"]):
                    tables.append((name_context, obj)); return
                if len(obj) > 0 and all(isinstance(v, dict) for v in obj.values()):
                    tables.append((name_context, obj)); return
                for k, v in obj.items(): hunt_tables(v, k)
            elif isinstance(obj, list) and obj:
                if isinstance(obj[0], dict) and any(k in obj[0] for k in ["rows", "dados", "linhas", "columns", "colunas"]):
                    for i, item in enumerate(obj):
                        hunt_tables(item, item.get("nome", item.get("titulo", item.get("tabela", f"{name_context} {i+1}"))))
                    return
                if isinstance(obj[0], (list, tuple, dict)):
                    tables.append((name_context, obj)); return

        hunt_tables(payload)
        out = []
        for name, raw_tab in tables:
            norm = _normalize_table_static(raw_tab, override_name=name)
            if norm and (norm["columns"] or norm["rows"]): out.append(norm)
        return out if out else []
    except Exception as e:
        print("Erro em ensure_payload_list:", e); return []

def _normalize_table_static(table, override_name=None):
    try:
        nome = override_name or "Tabela"
        cols = []
        raw_rows = []
        
        if isinstance(table, dict):
            nome = table.get("nome") or table.get("titulo") or table.get("tabela") or table.get("name") or override_name or "Tabela"
            
            for ck in ["columns", "colunas", "headings", "headers"]:
                if ck in table and isinstance(table[ck], list):
                    cols = [str(c) for c in table[ck]]; break
                    
            for rk in ["rows", "dados", "linhas", "data", "valores", "values"]:
                if rk in table and isinstance(table[rk], list):
                    raw_rows = table[rk]; break
                    
            if not raw_rows and all(isinstance(v, dict) for v in table.values()):
                df_keys = list(table.keys())
                all_x = set()
                for v in table.values(): all_x.update(v.keys())
                cols = ["HORARIO"] + df_keys
                for x in sorted(list(all_x)):
                    row = [x]
                    for k in df_keys: row.append(table[k].get(x, 0))
                    raw_rows.append(row)

        elif isinstance(table, list): raw_rows = table
            
        rows_out = []
        for linha in raw_rows:
            if isinstance(linha, dict):
                if "values" in linha and isinstance(linha["values"], dict):
                    vals = linha["values"]
                    if not cols: cols = list(vals.keys())
                    rows_out.append([vals.get(c) for c in cols])
                else:
                    if not cols: cols = list(linha.keys())
                    rows_out.append([linha.get(c) for c in cols])
            elif isinstance(linha, (list, tuple)):
                if not cols: cols = [f"COL_{i+1}" for i in range(len(linha))]
                rows_out.append(list(linha))
            else:
                if not cols: cols = ["VALOR"]
                rows_out.append([linha])

        headings = {c: c for c in cols}

        # Padroniza a coluna Horário para o gerador de Gráficos não falhar
        for i, c in enumerate(cols):
            if str(c).upper() in ["HORÁRIO", "HORARIO", "HORA"]:
                headings["HORARIO"] = cols[i]
                cols[i] = "HORARIO"

        return {"nome": nome, "columns": cols, "headings": headings, "rows": rows_out}
    except Exception as e: 
        return {"nome": override_name or "Tabela", "columns": [], "headings": {}, "rows": []}
    
def _create_combined_bar_figure_for_export(tabelas):
    try:
        # Pega especificamente a ÚLTIMA tabela do bloco (A Tabela Azul)
        t = tabelas[-1] if tabelas and len(tabelas) > 0 else None
        if not t or not t.get("rows"): return None

        cols = t.get("columns", [])
        headings = t.get("headings", {})
        rows = t.get("rows", [])
        
        idx_h = cols.index("HORARIO") if "HORARIO" in cols else 0
        horarios = [str(r[idx_h]) for r in rows]

        def heading_for(col):
            if isinstance(headings, dict): return headings.get(col, col)
            if isinstance(headings, list): return headings[cols.index(col)] if col in cols else col
            return col

        sentidos = [c for c in cols if c.upper() not in ["HORARIO", "TOTAL"] and "GERAL" not in c.upper()]
        if not sentidos: return None

        PALETTE = ["#FFD92E", "#58D68D", "#3498DB", "#E74C3C"]
        fig = Figure(figsize=(10, 3 * len(sentidos)), dpi=100)

        for i, sc in enumerate(sentidos, start=1):
            idx = cols.index(sc)
            valores = [safe_float(r[idx]) for r in rows]

            ax = fig.add_subplot(len(sentidos), 1, i)
            pos = list(range(len(valores)))
            cor = PALETTE[(i - 1) % len(PALETTE)]
            bars = ax.bar(pos, valores, color=cor)

            ax.set_title(f"{heading_for(sc)} - Valores por Horário")
            ax.set_xticks(pos)
            ax.set_xticklabels(horarios, rotation=45, ha="right", fontsize=8)
            ax.grid(axis="y", linestyle=":", alpha=0.6)

            for bar in bars:
                h = bar.get_height()
                txt = f"{int(h)}" if float(h).is_integer() else f"{h:.2f}"
                ax.annotate(txt, xy=(bar.get_x() + bar.get_width() / 2, h), xytext=(0, 3), textcoords="offset points", ha="center", va="bottom", fontsize=8)

        fig.tight_layout(pad=2.0)
        return fig
    except: return None


class RelatorioQuadroHorarioService:
    def __init__(self):
        self.repo = RelatorioQuadroHorarioRepository()

    def obter_linhas(self): return self.repo.obter_linhas()

    def abrir_documento(self, caminho):
        if not caminho or not os.path.exists(caminho):
            return False, "Arquivo não encontrado no diretório de rede."
        try:
            os.startfile(caminho)
            return True, "Abrindo documento..."
        except Exception as e:
            return False, f"Erro ao abrir o arquivo: {e}"

    def excluir_registro(self, tipo_doc, registro_id, motivo, excluido_por):
        return self.repo.excluir_registro(tipo_doc, registro_id, motivo, excluido_por)

    # Exportações Genéricas (Grade)
    def exportar_excel(self, tipo_doc, filtros, destino):
        dados = self.repo.buscar_dados_paginados(tipo_doc, filtros, limit=10000)
        if not dados: return False, "Nenhum dado encontrado para os filtros atuais."
        try:
            df = pd.DataFrame(dados)
            if 'id' in df.columns: df = df.drop(columns=['id'])
            if 'payload' in df.columns: df = df.drop(columns=['payload'])
            df.to_excel(destino, index=False)
            return True, "Relatório Excel gerado com sucesso."
        except Exception as e: return False, f"Erro ao gerar Excel: {e}"

    def exportar_pdf(self, tipo_doc, filtros, destino):
        dados = self.repo.buscar_dados_paginados(tipo_doc, filtros, limit=1000)
        if not dados: return False, "Nenhum dado encontrado."
        try:
            doc = SimpleDocTemplate(destino, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
            elementos = []
            estilos = getSampleStyleSheet()
            
            elementos.append(Paragraph(f"Relatório Gerencial - {tipo_doc} (Quadro de Horário)", estilos['Heading1']))
            elementos.append(Spacer(1, 15))

            if tipo_doc == "PESQUISA":
                cabecalho = ["ID", "Título", "Tipo de Pesquisa", "Início", "Fim", "Responsável", "Criação"]
                dados_tabela = [cabecalho]
                for d in dados:
                    dt_c = d.get('data_criacao').strftime("%d/%m/%Y") if d.get('data_criacao') else "-"
                    dt_i = d.get('data_inicio').strftime("%d/%m/%Y") if d.get('data_inicio') else "-"
                    dt_f = d.get('data_fim').strftime("%d/%m/%Y") if d.get('data_fim') else "-"
                    dados_tabela.append([str(d.get('id','')), str(d.get('titulo',''))[:40], str(d.get('tipo','')), dt_i, dt_f, str(d.get('responsavel',''))[:20], dt_c])
                col_widths = [40, 200, 150, 80, 80, 150, 80]
            else:
                cabecalho = ["Nº Parecer", "Processo", "Origem", "Assunto", "Decisão", "Solicitante", "Data Criação"]
                dados_tabela = [cabecalho]
                for d in dados:
                    dt = d.get('data_criacao').strftime("%d/%m/%Y") if d.get('data_criacao') else "-"
                    dados_tabela.append([
                        str(d.get('numero_completo','')), 
                        str(d.get('processo','')), 
                        str(d.get('origem','')), 
                        str(d.get('assunto',''))[:30], 
                        str(d.get('decisao','')), 
                        str(d.get('solicitante',''))[:20], 
                        dt
                    ])
                col_widths = [75, 95, 90, 180, 80, 150, 90] # Adaptado para folha A4 Paisagem

            tabela = Table(dados_tabela, colWidths=col_widths)
            tabela.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0F8C75")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F9F9F9")),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.silver),
            ]))
            
            elementos.append(tabela)
            doc.build(elementos)
            return True, "Relatório PDF gerado com sucesso!"
        except Exception as e: return False, f"Erro ao gerar PDF: {e}"

    # ==========================================
    # EXPORTAÇÕES AVANÇADAS (PESQUISA INDIVIDUAL)
    # ==========================================
    def exportar_pesquisa_excel(self, nome, tipo, payload, caminho):
        try:
            norm = ensure_payload_list(payload)
            if not norm: return False, "Não há dados estruturados para exportar."

            wb = Workbook()
            ws = wb.active
            ws.title = "Pesquisa"

            align_center = Alignment(horizontal="center", vertical="center")
            header_fill = PatternFill("solid", fgColor="C0C0C0")
            font_bold = Font(bold=True)

            def color_for_idx(idx):
                if idx < 3: return "ECF0F1"
                map_colors = ["F8D057", "3498DB", "5DADE2", "58D68D"] if tipo == "demanda" else ["F8D057", "58D68D", "3498DB", "E74C3C"]
                return map_colors[idx - 3] if idx - 3 < len(map_colors) else "FFFFFF"

            blocks = [norm[:3], norm[3:6]]
            block_starts = [1, 1 + max([len(t.get("rows", [])) for t in norm[:3]] + [0]) + 4]

            curr_col = 1
            for b_idx, block in enumerate(blocks):
                start_row = block_starts[b_idx]
                curr_col = 1
                for t_idx, tab in enumerate(block):
                    if not tab: continue
                    cols = tab.get("columns", [])
                    headings = tab.get("headings", {})
                    rows = tab.get("rows", [])
                    if not cols and rows: cols = [f"COL_{i}" for i in range(len(rows[0]))]
                    num_cols = max(1, len(cols))

                    ws.merge_cells(start_row=start_row, start_column=curr_col, end_row=start_row, end_column=curr_col + num_cols - 1)
                    tcell = ws.cell(row=start_row, column=curr_col, value=str(tab.get("nome", "Tabela")))
                    tcell.font = Font(bold=True, size=12); tcell.alignment = align_center
                    tcell.fill = PatternFill("solid", fgColor=color_for_idx(b_idx*3 + t_idx))

                    hr = start_row + 1
                    for j, col in enumerate(cols):
                        c = ws.cell(row=hr, column=curr_col + j, value=headings.get(col, col))
                        c.font = font_bold; c.alignment = align_center; c.fill = header_fill

                    rcur = hr + 1
                    for rvals in rows:
                        for j in range(num_cols):
                            val = rvals[j] if j < len(rvals) else ""
                            try:
                                num = float(str(val).replace(",", "."))
                                val = int(num) if num.is_integer() else num
                            except: pass
                            ws.cell(row=rcur, column=curr_col + j, value=val).alignment = align_center
                        rcur += 1

                    for j in range(num_cols):
                        ws.column_dimensions[get_column_letter(curr_col + j)].width = 15
                    curr_col += num_cols + 1

            fig = _create_combined_bar_figure_for_export(norm[3:6])
            if fig:
                temp_dir = tempfile.mkdtemp()
                img_path = os.path.join(temp_dir, "chart.png")
                fig.savefig(img_path, bbox_inches="tight")
                plt.close(fig)
                
                row_img = block_starts[1] + max([len(t.get("rows", [])) for t in norm[3:6]] + [0]) + 3
                ws.add_image(XLImage(img_path), f"A{row_img}")

            wb.save(caminho)
            return True, "Exportação Avançada em Excel concluída com sucesso!"
        except Exception as e: return False, f"Falha na exportação Excel: {e}"

    def exportar_pesquisa_pdf(self, nome_pesquisa, tipo_pesquisa, payload, destino):
        try:
            # --- BLINDAGEM JSON ---
            if isinstance(payload, str):
                import json
                try: payload = json.loads(payload)
                except: pass
            # ----------------------

            if not payload or not isinstance(payload, dict) or 'tabelas' not in payload:
                return False, "O payload da pesquisa está vazio ou inválido."

            # Folha A4 Paisagem (Landscape) para caber as 3 tabelas lado a lado
            doc = SimpleDocTemplate(destino, pagesize=landscape(A4), rightMargin=15, leftMargin=15, topMargin=15, bottomMargin=15)
            elementos = []
            estilos = getSampleStyleSheet()
            
            # Cabeçalho do PDF
            elementos.append(Paragraph(f"Relatório Analítico de Pesquisa: {nome_pesquisa}", estilos['Heading1']))
            elementos.append(Paragraph(f"Tipo: {str(tipo_pesquisa).upper()}", estilos['Heading3']))
            elementos.append(Spacer(1, 15))

            tabelas_payload = payload.get('tabelas', [])
            datas_pesquisa = payload.get('datas', [])
            
            # =========================================================
            # BLOCO 1: As 3 Tabelas Principais (Relatórios 1, 2 e 3)
            # =========================================================
            celulas_horizontais_superiores = []
            for idx, tab in enumerate(tabelas_payload[:3]): 
                nome_tab = tab.get('tabela', f'Relatório {idx+1}')
                colunas = tab.get('colunas', [])
                linhas = tab.get('linhas', [])
                
                # MÁGICA 1: Substitui a palavra "Horário" pela Data correspondente do Banco
                if colunas and idx < len(datas_pesquisa):
                    colunas[0] = str(datas_pesquisa[idx])
                    
                tabela_dados = [colunas] + linhas
                cor_cabecalho = "#3498DB" if tipo_pesquisa.lower() == "demanda" else "#0F8C75"
                
                t = Table(tabela_dados)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(cor_cabecalho)),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.silver),
                ]))
                
                titulo_tab = Paragraph(f"<para align='center'><b>{nome_tab}</b></para>", estilos['Normal'])
                celulas_horizontais_superiores.append([titulo_tab, Spacer(1, 5), t])

            # Agrupa as 3 tabelas numa "Tabela Mestre Invisível" para ficarem lado a lado
            if celulas_horizontais_superiores:
                largura_coluna = 800 / len(celulas_horizontais_superiores)
                tabela_mestre_sup = Table([celulas_horizontais_superiores], colWidths=[largura_coluna]*len(celulas_horizontais_superiores))
                tabela_mestre_sup.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 5),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ]))
                elementos.append(tabela_mestre_sup)
                
            # =========================================================
            # BLOCO 2: As Tabelas de Baixo (Média, Quadro, Diferença)
            # =========================================================
            celulas_horizontais_inferiores = []
            for idx, tab in enumerate(tabelas_payload[3:6]):
                nome_tab = tab.get('tabela', f'Resumo {idx+1}')
                colunas = tab.get('colunas', [])
                linhas = tab.get('linhas', [])
                
                tabela_dados = [colunas] + linhas
                
                # MÁGICA 2: Mapeia as cores idênticas às da interface do usuário
                if "Média" in nome_tab: 
                    cor = "#F8D057" # Amarelo
                elif "Quadro" in nome_tab or "Passageiro" in nome_tab: 
                    cor = "#96D37A" # Verde
                else: 
                    cor = "#70ADE7" # Azul
                
                t = Table(tabela_dados)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(cor)),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.silver),
                ]))
                
                titulo_tab = Paragraph(f"<para align='center'><b>{nome_tab}</b></para>", estilos['Normal'])
                celulas_horizontais_inferiores.append([titulo_tab, Spacer(1, 5), t])
                
            if celulas_horizontais_inferiores:
                elementos.append(Spacer(1, 15)) # Espaço entre a linha de cima e a de baixo
                tabela_mestre_inf = Table([celulas_horizontais_inferiores], colWidths=[largura_coluna]*len(celulas_horizontais_inferiores))
                tabela_mestre_inf.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 5),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ]))
                elementos.append(tabela_mestre_inf)

            # Renderiza e Salva o PDF
            doc.build(elementos)
            return True, "Relatório exportado em PDF com sucesso!"
            
        except Exception as e:
            return False, f"Erro ao gerar o PDF: {e}"